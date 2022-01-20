

import time
import sys
import os
from openpyxl import load_workbook

def logo():
    print("""
        
  ______ __   __ _____  ______  _        _______  ____      _____  _    _               
 |  ____|\ \ / // ____||  ____|| |      |__   __|/ __ \    / ____|| |  (_)              
 | |__    \ V /| |     | |__   | |         | |  | |  | |  | (___  | |_  _  _ __    __ _ 
 |  __|    > < | |     |  __|  | |         | |  | |  | |   \___ \ | __|| || '_ \  / _` |
 | |____  / . \| |____ | |____ | |____     | |  | |__| |   ____) || |_ | || | | || (_| |
 |______|/_/ \_\\\_____||______||______|    |_|   \____/   |_____/  \__||_||_| |_| \__, |
                                                                                   __/ |
                                                                                  |___/ 
        
    """)

def __help():
    print("æ­¤è„šæœ¬ç”¨æ¥å¸®åŠ©å¼€å‘è€…ä»Žexcelç”Ÿæˆå›½é™…åŒ–æ–‡ä»¶ï¼ˆiOSã€Androidï¼‰")
    print("é¦–å…ˆéœ€è¦å®‰è£…`openpyxl`åº“ï¼Œå¯ä»¥ä½¿ç”¨`pip3 install openpyxl`å‘½ä»¤å®‰è£…")
    print("-------------å‚æ•°--------------\n")
    print("å‚æ•° `-h` æˆ– `--help`    æŸ¥çœ‹å¸®åŠ©.")
    print("å‚æ•° `-ep` æˆ– `--excelPath`         excelæ–‡ä»¶çš„ç›®å½•ï¼Œé»˜è®¤ä¸º`~/Desktop/international_test.excel`")
    print("å‚æ•° `-sp` æˆ– `--stringFilePath`    å¯¼å‡ºå›½é™…åŒ–æ–‡ä»¶çš„ç›®å½•ï¼Œé»˜è®¤ä¸º`~/Desktop/InternationalStrings/`")
    print("å‚æ•° `-kc` æˆ– `--keyColumn`         æŒ‡å®škeyåœ¨excelä¸­çš„åˆ—ï¼Œé»˜è®¤ä¸º`B`")
    print("å‚æ•° `-vc` æˆ– `--valueColumn`       æŒ‡å®švalueåœ¨excelä¸­çš„åˆ—ï¼Œé»˜è®¤ä¸º`[C]`ï¼Œå¯ä»¥ä¼ å…¥æ•°ç»„(ä»¥`,`åˆ†å‰²)ï¼Œå¯ä»¥æ ¹æ®ä¼ å…¥æ•°ç»„ç”Ÿæˆä¸åŒçš„å›½é™…åŒ–æ–‡ä»¶")
    print("å‚æ•° `-is` æˆ– `--ignoreSheets`      æŒ‡å®šå¿½ç•¥çš„sheetï¼Œå¯ä»¥ä¼ å…¥æ•°ç»„ï¼Œ(ä»¥`,`åˆ†å‰²)")
    print("-------------------------------\n")


# æ‰“å¼€excel
def loadExcel(excelFilePath):
    wb = load_workbook(filename=excelFilePath)
    return wb

# è¿‡æ»¤ä¸éœ€è¦çš„sheet
def filterSheetNames(filtration,sheetNames):
    newNames = sheetNames
    for filter in filtration:
        if filter in newNames:
            newNames.remove(filter)
    return newNames

# è¯»å–sheetå¹¶å¤„ç†æ‰€æœ‰æ–‡æ¡ˆï¼Œ å¹¶å°†å¤„ç†åŽçš„æ–‡æ¡ˆä¿å­˜åˆ°`allStringDict`ï¼Œ å°†æ‰€æœ‰keyä¿å­˜åˆ°`keys`
def processSheetStringList(sheetName, wb, valueColumn, needWriteKeys):
    tmpSheet = wb[sheetName] # å–å‡ºæŒ‡å®šçš„sheet
    tmpSheetMaxRow = tmpSheet.max_row
    valueName = tmpSheet["%s1"%valueColumn].value
    for row in range(tmpSheetMaxRow):
        keyIndex = "B%s"%(row + 2)  # ä»Žç¬¬äºŒè¡Œå¼€å§‹è¯»å–
        valueIndex = "%s%s"%(valueColumn,row + 2)
        key = "%s"%tmpSheet[keyIndex].value
        value = "%s"%tmpSheet[valueIndex].value
        if key != "None" and len(key) > 0 and value != None and len(value) > 0:   # keyå’Œvalueéƒ½ä¸ä¸ºç©ºæ—¶å†™å…¥æ–‡ä»¶
            if needWriteKeys == True:
                if (key in keys) == False:
                    keys.append(key)    # æ·»åŠ æ–°key
                else:
                    keys.remove(key)    # åˆ æŽ‰åŽŸæ¥çš„key
                    keys.append(key)    # å°†keyæ·»åŠ åˆ°å°¾éƒ¨
            allStringDict[key] = value
    return valueName

# å°†å¤„ç†å¥½çš„æ–‡ä»¶å†™å…¥å›½é™…åŒ–æ–‡ä»¶
def writeAllStringToIntenationalFile(file):
    for key in keys:
        value = allStringDict[key]
        if len(value) > 0:
            file.write('\"%s\" = \"%s\";\n' %(key, value))

# åˆ›å»ºæ–‡ä»¶å¹¶æ·»åŠ æ–‡ä»¶å¤´
def create_iOS_InitializeStringFile(path):
    if path != None and len(path) > 0:
        if os.path.exists(path) == False:
            os.makedirs(path)
        tmpStringFile = open('%s/Localizable.strings'%path, 'w')

    if tmpStringFile != None:
        tmpStringFile.write("""
/* 
  Localizable.strings
  YourProjectName

  Created by walker on %s.
  Copyright Â© %s YourProjectName. All rights reserved.
*/


"""%(time.strftime("%Y/%m/%d %H:%M", time.localtime()), time.strftime("%Y", time.localtime())))
    return tmpStringFile

# å†™å…¥å›½é™…åŒ–æ–‡ä»¶
def writeInternationalStringToFile(filePath):
    file = create_iOS_InitializeStringFile(filePath)
    writeAllStringToIntenationalFile(file)
    file.close()
    print(' ðŸ¤– [%s/Localizable.strings]å†™å…¥å®Œæˆï¼\n' % filePath)

# è½¬æ¢excelä¸­æŒ‡å®šçš„valueä¸ºå›½é™…åŒ–æ–‡ä»¶
def convertExcelToString(valueColumn):
    valueName = ""
    for sheetName in needProcessSheetNames: # éåŽ†æ‰€æœ‰sheetï¼Œå¹¶å°†æ‰€æœ‰çš„key-valueç¼“å­˜èµ·æ¥
        tmpValueName = processSheetStringList(sheetName, excel, valueColumn, True)
        if tmpValueName != None and len(tmpValueName) > 0 and len(valueName) <= 0:
            valueName = tmpValueName
    writeInternationalStringToFile("%s/%s" % (outPath, valueName))

# main å‡½æ•°
if __name__ == "__main__":
    # æ˜¾ç¤ºlogo
    logo()
    # è¾“å‡ºä¼ å…¥å‚æ•°
    print('ä¼ å…¥å‚æ•°ä¸ºï¼š%s\n'%sys.argv)
    # åˆå§‹åŒ–å˜é‡
    keys = []   # keyå­˜çš„æ•°ç»„ï¼Œç”¨æ¥ä¿å­˜æ‰€æœ‰çš„key
    allStringDict = {}  # key valueå­˜çš„å­—å…¸
    excelPath = '/Users/apple/Desktop/international_test.xlsx'
    outPath = 'InternationalStrings'
    keyColumn = 'B'
    valueColumns = ['C'] # é»˜è®¤è½¬æ¢çš„å›½é™…åŒ–åˆ—åç§°
    ignoreSheets = ['what\'s new','Backend']

    # åˆ†å‰²å¹¶å¤„ç†ä¼ å…¥çš„å‚æ•°
    for tmpArg in sys.argv:
        if '-h' in tmpArg or '--help' in tmpArg:
            __help()
            exit(0)
        elif '-ep=' in tmpArg:
            excelPath = tmpArg.replace('-ep=', '')
        elif '--excelPath=' in tmpArg:
            excelPath = tmpArg.replace('--excelPath=', '')
        elif '-sp=' in tmpArg:
            outPath = tmpArg.replace('-sp=', '')
        elif '--stringFilePath=' in tmpArg:
            outPath = tmpArg.replace('--stringFilePath=', '')
        elif '-kc=' in tmpArg:
            keyColumn = tmpArg.replace('-kc=', '').upper()
        elif '--keyColumn=' in tmpArg:
            keyColumn = tmpArg.replace('--keyColumn=', '').upper()
        elif '-vc=' in tmpArg:
            valueColumns = tmpArg.replace('-vc=', '').upper().split(',')
        elif '--valueColumn=' in tmpArg:
            valueColumns = tmpArg.replace('--valueColumn=', '').upper().split(',')
        elif '-is=' in tmpArg:
            ignoreSheets = tmpArg.replace('-is=', '').upper().split(',')
        elif '--ignoreSheets=' in tmpArg:
            ignoreSheets = tmpArg.replace('--ignoreSheets=', '').upper().split(',')

    #print(excelPath, outPath, keyColumn, valueColumns, ignoreSheets)

    # è¯»å–excelæ–‡ä»¶
    excel = loadExcel(excelPath)

    # è¿‡æ»¤ä¸éœ€è¦çš„sheet
    needProcessSheetNames = filterSheetNames(ignoreSheets, excel.sheetnames)

    # éåŽ†è½¬æ¢æ‰€æœ‰å›½é™…åŒ–æ–‡æ¡ˆ
    for vc in valueColumns:
        convertExcelToString(vc)    # é€åˆ—è½¬æ¢stringæ–‡ä»¶
    print('âœ… å¤„ç†å®Œæˆï¼')